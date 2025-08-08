import openpyxl as excel

e = excel.load_workbook("C:\\Users\\dilli\\Downloads\\Network Applications (1) (version 1).xlsb.xlsx")
active_sheet = e.active
cells = active_sheet.cell(row=1, column=1)
print(cells.value)
active_sheet.cell(row=27,column=1).value = "26"
print(active_sheet.cell(row=27,column=1).value)
print(active_sheet.max_row)
print(active_sheet.max_column)

# for r in active_sheet.iter_rows():
#     for cell in r:
#         print(cell.value, end="\t")
#     print()
# empty dict
dictionary = dict()
for x in range(1, active_sheet.max_row + 1):
     if active_sheet.cell(row=x,column=1).value == 5:
        for y in range(1, active_sheet.max_column + 1):
            print(active_sheet.cell(row=x,column=y).value, end="\t")
            dictionary[active_sheet.cell(row=1, column=y).value  ] = active_sheet.cell(row=x,column=y).value
        print()

print(dictionary)